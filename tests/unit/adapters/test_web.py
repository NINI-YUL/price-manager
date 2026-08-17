from __future__ import annotations

import hashlib
from collections import Counter
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from src.adapters import WebAdapterConfig, WebPriceAdapter
from src.models import Channel, ImportTaskStatus

TIERS = frozenset({Decimal("0.99"), Decimal("1.99")})


def test_valid_workbook_ignores_income_and_color_without_modifying_source(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "web-valid.xlsx"
    _write_workbook(
        workbook_path,
        headers=(("乌克兰", "USD"), ("阿拉伯联合酋长国", "AED")),
        tiers=(0.99, 1.99),
        prices=((0.99, 4.99), (1.99, 9.99)),
        incomes=((0.7, 3.2), (1.4, 6.4)),
        yellow_cell="D3",
    )
    before = _sha256(workbook_path)

    result = _adapter().parse(workbook_path)

    assert _sha256(workbook_path) == before == result.source_sha256
    assert result.status is ImportTaskStatus.CHECKING
    assert result.channel is Channel.WEB
    assert result.selected_sheet == "输出表"
    assert result.issues == ()
    assert result.statistics.source_row_count == 2
    assert result.statistics.product_count == 0
    assert result.statistics.price_cell_count == 4
    assert result.statistics.accepted_record_count == 4
    assert result.statistics.country_count == 2
    assert result.statistics.currency_count == 2
    assert result.statistics.tier_count == 2
    ua = next(
        record
        for record in result.records
        if record.country_code == "UA" and record.usd_tier == Decimal("0.99")
    )
    assert ua.currency == "USD"
    assert ua.local_price == Decimal("0.99")
    assert ua.product_id is None
    assert ua.adjustment_mode is None
    assert (ua.source_sheet, ua.source_row, ua.source_column) == ("输出表", 3, "B")


@pytest.mark.parametrize("suffix", [".xls", ".csv", ".txt"])
def test_unsupported_file_type_is_fatal(tmp_path: Path, suffix: str) -> None:
    source = tmp_path / f"web{suffix}"
    source.write_bytes(b"not an xlsx")

    result = _adapter().parse(source)

    assert result.status is ImportTaskStatus.FAILED
    assert _codes(result) == Counter({"W001": 1})


def test_corrupt_and_oversized_workbooks_are_fatal(tmp_path: Path) -> None:
    corrupt = tmp_path / "corrupt.xlsx"
    corrupt.write_bytes(b"not an xlsx package")
    valid = tmp_path / "oversized.xlsx"
    _write_workbook(valid)

    corrupt_result = _adapter().parse(corrupt)
    oversized_result = _adapter(max_file_size_bytes=1).parse(valid)

    assert corrupt_result.status is ImportTaskStatus.FAILED
    assert oversized_result.status is ImportTaskStatus.FAILED
    assert _codes(corrupt_result) == Counter({"W001": 1})
    assert _codes(oversized_result) == Counter({"W001": 1})


def test_missing_ambiguous_and_malformed_layouts_are_fatal(tmp_path: Path) -> None:
    missing = tmp_path / "missing-layout.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "not a web table"
    workbook.save(missing)

    ambiguous = tmp_path / "ambiguous.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "输出表"
    _populate_sheet(first)
    second = workbook.create_sheet("另一个输出表")
    _populate_sheet(second)
    workbook.save(ambiguous)

    malformed = tmp_path / "malformed.xlsx"
    _write_workbook(malformed)
    workbook = _load_editable(malformed)
    workbook["输出表"]["C2"] = "成本"
    workbook.save(malformed)

    assert _codes(_adapter().parse(missing)) == Counter({"W002": 1})
    assert _codes(_adapter().parse(ambiguous)) == Counter({"W002": 1})
    malformed_result = _adapter().parse(malformed)
    assert malformed_result.status is ImportTaskStatus.FAILED
    assert _codes(malformed_result) == Counter({"W002": 1})


def test_unknown_country_and_currency_are_located(tmp_path: Path) -> None:
    workbook_path = tmp_path / "bad-columns.xlsx"
    _write_workbook(
        workbook_path,
        headers=(("未知地区", "USD"), ("乌克兰", "ZZZ")),
    )

    result = _adapter().parse(workbook_path)

    assert result.status is ImportTaskStatus.CHECKING
    assert _codes(result) == Counter({"W003": 1, "W004": 1})
    assert {(issue.code, issue.source_column) for issue in result.issues} == {
        ("W003", "B"),
        ("W004", "D"),
    }


def test_bad_tier_blank_price_and_missing_tier_are_reported(tmp_path: Path) -> None:
    workbook_path = tmp_path / "bad-data.xlsx"
    _write_workbook(
        workbook_path,
        headers=(("乌克兰", "USD"),),
        tiers=(0.99, 3.99),
        prices=((None,), (4.0,)),
    )

    result = _adapter().parse(workbook_path)

    assert _codes(result) == Counter({"W005": 1, "W006": 1, "W008": 1})
    assert result.statistics.accepted_record_count == 0
    assert {(issue.code, issue.source_column) for issue in result.issues} >= {
        ("W005", "A"),
        ("W006", "B"),
    }


def test_formula_without_cached_value_is_reported(tmp_path: Path) -> None:
    workbook_path = tmp_path / "formula.xlsx"
    _write_workbook(workbook_path, headers=(("乌克兰", "USD"),))
    workbook = _load_editable(workbook_path)
    workbook["输出表"]["B3"] = "=1+1"
    workbook.save(workbook_path)

    result = _adapter().parse(workbook_path)

    assert _codes(result) == Counter({"W006": 1})
    issue = result.issues[0]
    assert (issue.source_row, issue.source_column) == (3, "B")


def test_identical_duplicate_is_warned_and_conflict_is_blocking(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "duplicates.xlsx"
    _write_workbook(
        workbook_path,
        headers=(("乌克兰", "USD"),),
        tiers=(0.99, 0.99, 1.99, 1.99),
        prices=((1.0,), (1.0,), (2.0,), (3.0,)),
    )

    result = _adapter().parse(workbook_path)

    assert _codes(result) == Counter({"W102": 1, "W007": 1})
    assert result.statistics.accepted_record_count == 2
    assert result.statistics.duplicate_count == 1


def test_duplicate_country_column_is_blocking(tmp_path: Path) -> None:
    workbook_path = tmp_path / "duplicate-country.xlsx"
    _write_workbook(
        workbook_path,
        headers=(("乌克兰", "USD"), ("乌克兰", "USD")),
    )

    result = _adapter().parse(workbook_path)

    assert _codes(result) == Counter({"W007": 1})
    assert result.statistics.accepted_record_count == 2


def _adapter(*, max_file_size_bytes: int = 50 * 1024 * 1024) -> WebPriceAdapter:
    return WebPriceAdapter(
        WebAdapterConfig(
            country_names={
                "乌克兰": "UA",
                "阿拉伯联合酋长国": "AE",
                "捷克共和国": "CZ",
            },
            supported_currency_codes=frozenset({"USD", "AED", "CZK"}),
            configured_tiers=TIERS,
            max_file_size_bytes=max_file_size_bytes,
        )
    )


def _write_workbook(
    path: Path,
    *,
    headers: tuple[tuple[str, str], ...] = (("乌克兰", "USD"),),
    tiers: tuple[float, ...] = (0.99, 1.99),
    prices: tuple[tuple[float | str | None, ...], ...] | None = None,
    incomes: tuple[tuple[float | str | None, ...], ...] | None = None,
    yellow_cell: str | None = None,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "输出表"
    _populate_sheet(sheet, headers=headers, tiers=tiers, prices=prices, incomes=incomes)
    if yellow_cell is not None:
        sheet[yellow_cell].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    workbook.save(path)


def _populate_sheet(
    sheet,
    *,
    headers: tuple[tuple[str, str], ...] = (("乌克兰", "USD"),),
    tiers: tuple[float, ...] = (0.99, 1.99),
    prices: tuple[tuple[float | str | None, ...], ...] | None = None,
    incomes: tuple[tuple[float | str | None, ...], ...] | None = None,
) -> None:
    sheet.cell(1, 1, "积分")
    sheet.cell(2, 1, "积分")
    for index, (country, currency) in enumerate(headers):
        price_column = 2 + index * 2
        sheet.cell(1, price_column, f"{country}({currency})")
        sheet.cell(2, price_column, "价格")
        sheet.cell(2, price_column + 1, "收入")
    price_rows = prices or tuple(
        tuple(tier * (index + 1) for index in range(len(headers))) for tier in tiers
    )
    income_rows = incomes or tuple(tuple(None for _ in headers) for _ in tiers)
    for row_offset, tier in enumerate(tiers, start=3):
        sheet.cell(row_offset, 1, tier)
        for country_index in range(len(headers)):
            column = 2 + country_index * 2
            sheet.cell(row_offset, column, price_rows[row_offset - 3][country_index])
            sheet.cell(
                row_offset,
                column + 1,
                income_rows[row_offset - 3][country_index],
            )


def _load_editable(path: Path):
    from openpyxl import load_workbook

    return load_workbook(path)


def _codes(result) -> Counter[str]:
    return Counter(issue.code for issue in result.issues)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()
