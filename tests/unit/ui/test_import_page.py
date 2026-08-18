from __future__ import annotations

import os
from decimal import Decimal
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import Workbook
from PySide6.QtCore import QEventLoop, QTimer

from src.database.connection import database_session
from src.database.repositories import ImportTaskRepository
from src.database.seed import EXPECTED_PRICE_TIERS, seed_database
from src.main import create_application, create_main_window
from src.models import (
    Channel,
    ImportIssue,
    ImportPreview,
    ImportStatistics,
    ImportTaskStatus,
    IssueSeverity,
    StandardPrice,
)
from src.ui import ApplicationShell, ImportPage
from src.utils.source_hash import file_sha256


def test_page_channel_modes_preview_filters_and_confirmation_gate(tmp_path: Path) -> None:
    _application()
    database_path = tmp_path / "ui.db"
    seed_database(database_path)
    page = ImportPage(database_path, archives_root=tmp_path / "archives")

    assert page.selected_channel is Channel.GOOGLE
    assert page.source_selection_mode == "file"
    page.channel_combo.setCurrentIndex(1)
    assert page.selected_channel is Channel.IOS
    assert page.source_selection_mode == "directory"
    assert page.browse_button.text() == "选择目录"
    page.channel_combo.setCurrentIndex(2)

    preview = _preview(issues=(ImportIssue("W102", IssueSeverity.WARNING, "duplicate removed"),))
    page.show_preview(preview, "TASK-UI")

    assert page.confirm_button.isEnabled()
    assert page.price_table.rowCount() == 2
    assert page.issue_table.rowCount() == 1
    assert "接受价格：2" in page.summary_label.text()
    page.country_filter.setCurrentIndex(page.country_filter.findData("JP"))
    assert page.price_table.rowCount() == 1
    assert page.price_table.item(0, 0).text() == "JP"
    page.severity_filter.setCurrentIndex(page.severity_filter.findData("ERROR"))
    assert page.issue_table.rowCount() == 0

    blocking = _preview(issues=(ImportIssue("W006", IssueSeverity.ERROR, "bad price"),))
    page.show_preview(blocking, "TASK-BLOCKED")
    assert not page.confirm_button.isEnabled()
    page.close()


def test_page_parses_web_workbook_in_background(tmp_path: Path) -> None:
    application = _application()
    database_path = tmp_path / "background.db"
    seed_database(database_path)
    workbook_path = tmp_path / "web.xlsx"
    _write_web_workbook(workbook_path)
    page = ImportPage(database_path, archives_root=tmp_path / "archives")
    page.channel_combo.setCurrentIndex(2)
    page.set_source_path(workbook_path)

    loop = QEventLoop()
    timeout = {"fired": False}

    def finish(_preview, _task_id) -> None:
        QTimer.singleShot(50, loop.quit)

    def fail(_message: str) -> None:
        QTimer.singleShot(50, loop.quit)

    def time_out() -> None:
        timeout["fired"] = True
        loop.quit()

    page.preview_loaded.connect(finish)
    page.parse_failed.connect(fail)
    QTimer.singleShot(5000, time_out)
    page.start_parse()
    loop.exec()
    application.processEvents()

    assert not timeout["fired"]
    assert page.current_preview is not None
    assert page.current_preview.status is ImportTaskStatus.CHECKING
    assert page.current_preview.statistics.accepted_record_count == 14
    assert page.confirm_button.isEnabled()
    page.close()


def test_main_window_uses_navigation_shell_with_import_default(tmp_path: Path) -> None:
    _application()
    database_path = tmp_path / "window.db"
    seed_database(database_path)
    window = create_main_window(
        database_path,
        archives_path=tmp_path / "archives",
    )

    shell = window.centralWidget()
    assert isinstance(shell, ApplicationShell)
    assert isinstance(shell.import_page, ImportPage)
    assert shell.stack.currentWidget() is shell.import_page
    assert shell.import_navigation.isChecked()
    assert window.minimumWidth() == 1180
    window.close()


def test_successful_ui_confirmation_disables_repeat_action(tmp_path: Path, monkeypatch) -> None:
    _application()
    database_path = tmp_path / "confirm-ui.db"
    seed_database(database_path)
    source = tmp_path / "web.xlsx"
    source.write_bytes(b"ui-confirmation")
    preview = _preview(issues=())
    preview = ImportPreview(
        channel=preview.channel,
        source_path=str(source.resolve()),
        source_sha256=file_sha256(source),
        selected_sheet=preview.selected_sheet,
        status=preview.status,
        records=preview.records,
        issues=preview.issues,
        statistics=preview.statistics,
    )
    with database_session(database_path) as connection:
        tasks = ImportTaskRepository(connection)
        tasks.create(
            task_id="TASK-UI-CONFIRM",
            channel=Channel.WEB,
            file_path=preview.source_path,
            created_time="2026-08-17T16:00:00+08:00",
        )
        tasks.update_result(
            task_id="TASK-UI-CONFIRM",
            status=ImportTaskStatus.CHECKING,
            error_count=0,
            warning_count=0,
            completed_time=None,
            error_message=None,
        )
    monkeypatch.setattr("src.ui.import_page.QMessageBox.information", lambda *args: None)
    page = ImportPage(database_path, archives_root=tmp_path / "archives")
    page.show_preview(preview, "TASK-UI-CONFIRM")

    page.confirm_button.click()

    assert not page.confirm_button.isEnabled()
    assert "入库成功" in page.status_label.text()
    with database_session(database_path) as connection:
        task = ImportTaskRepository(connection).get("TASK-UI-CONFIRM")
        assert task["status"] == "SUCCESS"
        assert task["version_id"].startswith("WEB_V")
    page.close()


def _application():
    return create_application(["price-manager-test"])


def _preview(*, issues: tuple[ImportIssue, ...]) -> ImportPreview:
    records = (
        _record("US", "USD", "0.99"),
        _record("JP", "JPY", "150"),
    )
    return ImportPreview(
        channel=Channel.WEB,
        source_path=str(Path("web.xlsx").resolve()),
        source_sha256="a" * 64,
        selected_sheet="Sheet1",
        status=ImportTaskStatus.CHECKING,
        records=records,
        issues=issues,
        statistics=ImportStatistics(
            accepted_record_count=2,
            country_count=2,
            currency_count=2,
            tier_count=1,
            error_count=sum(issue.severity is IssueSeverity.ERROR for issue in issues),
            warning_count=sum(issue.severity is IssueSeverity.WARNING for issue in issues),
        ),
    )


def _record(country: str, currency: str, price: str) -> StandardPrice:
    return StandardPrice(
        channel=Channel.WEB,
        country_code=country,
        usd_tier=Decimal("0.99"),
        currency=currency,
        local_price=Decimal(price),
        product_id=None,
        source_sheet="Sheet1",
        source_row=3,
        source_column="B",
    )


def _write_web_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "输出表"
    sheet["A1"] = "积分"
    sheet["A2"] = "积分"
    sheet["B1"] = "美国(USD)"
    sheet["B2"] = "价格"
    sheet["C2"] = "收入"
    for row, tier in enumerate(EXPECTED_PRICE_TIERS, start=3):
        sheet.cell(row, 1, float(tier))
        sheet.cell(row, 2, float(tier))
    workbook.save(path)
