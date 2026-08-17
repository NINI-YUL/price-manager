from __future__ import annotations

from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from src.database.connection import database_session
from src.database.repositories import ReferenceDataRepository
from src.database.schema import initialize_database
from src.models import ImportTaskStatus
from src.services import WebImportService
from src.services.web_import import load_web_country_aliases

FIXED_TIME = datetime(2026, 8, 17, 15, 0, tzinfo=timezone(timedelta(hours=8)))


def test_service_accepts_header_currency_alias_and_writes_no_formal_prices(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "web-import.db"
    aliases_path = _initialize_reference_data(database_path, tmp_path)
    source = _write_web_workbook(tmp_path / "web.xlsx")

    result = WebImportService(
        database_path, aliases_path=aliases_path, clock=lambda: FIXED_TIME
    ).preview(source, task_id="WEB_TASK_VALID")

    assert result.status is ImportTaskStatus.CHECKING
    assert result.issues == ()
    assert result.statistics.accepted_record_count == 4
    assert {record.country_code for record in result.records} == {"UA", "CZ"}
    assert (
        next(record for record in result.records if record.country_code == "UA").currency == "USD"
    )
    with database_session(database_path) as connection:
        task = connection.execute(
            "SELECT * FROM import_tasks WHERE task_id = 'WEB_TASK_VALID'"
        ).fetchone()
        assert task["channel"] == "WEB"
        assert task["status"] == "CHECKING"
        assert task["error_count"] == 0
        assert task["completed_time"] is None
        assert connection.execute("SELECT COUNT(*) FROM channel_prices").fetchone()[0] == 0
        assert connection.execute("SELECT COUNT(*) FROM price_versions").fetchone()[0] == 0


def test_fatal_parse_updates_task_to_failed(tmp_path: Path) -> None:
    database_path = tmp_path / "web-fatal.db"
    aliases_path = _initialize_reference_data(database_path, tmp_path)
    corrupt = tmp_path / "corrupt.xlsx"
    corrupt.write_bytes(b"not an xlsx package")

    result = WebImportService(
        database_path, aliases_path=aliases_path, clock=lambda: FIXED_TIME
    ).preview(corrupt, task_id="WEB_TASK_FATAL")

    assert result.status is ImportTaskStatus.FAILED
    with database_session(database_path) as connection:
        task = connection.execute(
            "SELECT * FROM import_tasks WHERE task_id = 'WEB_TASK_FATAL'"
        ).fetchone()
        assert task["status"] == "FAILED"
        assert task["error_count"] == 1
        assert task["completed_time"] == FIXED_TIME.isoformat()
        assert task["error_message"]


def test_unexpected_parser_error_is_recorded_and_propagated(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    database_path = tmp_path / "web-unexpected.db"
    aliases_path = _initialize_reference_data(database_path, tmp_path)
    source = _write_web_workbook(tmp_path / "web.xlsx")

    def fail_parse(*_args, **_kwargs):
        raise RuntimeError("synthetic web parser failure")

    monkeypatch.setattr("src.services.web_import.WebPriceAdapter.parse", fail_parse)
    with pytest.raises(RuntimeError, match="synthetic web parser failure"):
        WebImportService(
            database_path, aliases_path=aliases_path, clock=lambda: FIXED_TIME
        ).preview(source, task_id="WEB_TASK_UNEXPECTED")

    with database_session(database_path) as connection:
        task = connection.execute(
            "SELECT * FROM import_tasks WHERE task_id = 'WEB_TASK_UNEXPECTED'"
        ).fetchone()
        assert task["status"] == "FAILED"
        assert "synthetic web parser failure" in task["error_message"]


def test_reference_data_and_alias_config_are_validated_before_task(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "empty.db"
    initialize_database(database_path)
    aliases_path = tmp_path / "aliases.csv"
    aliases_path.write_text("alias,country_code\n捷克共和国,CZ\n", encoding="utf-8")

    with pytest.raises(RuntimeError, match="P1-003 reference data"):
        WebImportService(database_path, aliases_path=aliases_path).preview(
            tmp_path / "web.xlsx", task_id="WEB_TASK_EMPTY"
        )

    with pytest.raises(ValueError, match="invalid web country alias"):
        load_web_country_aliases(aliases_path, {"UA"})
    with database_session(database_path) as connection:
        assert connection.execute("SELECT COUNT(*) FROM import_tasks").fetchone()[0] == 0


def _initialize_reference_data(database_path: Path, tmp_path: Path) -> Path:
    initialize_database(database_path)
    with database_session(database_path) as connection:
        repository = ReferenceDataRepository(connection)
        repository.add_country(
            country_code="UA",
            name_cn="乌克兰",
            name_en="Ukraine",
            default_currency="UAH",
        )
        repository.add_country(
            country_code="CZ",
            name_cn="捷克",
            name_en="Czechia",
            default_currency="CZK",
        )
        repository.add_country(
            country_code="US",
            name_cn="美国",
            name_en="United States",
            default_currency="USD",
        )
        repository.add_price_tier(Decimal("0.99"))
        repository.add_price_tier(Decimal("1.99"))
    aliases_path = tmp_path / "web-aliases.csv"
    aliases_path.write_text("alias,country_code\n捷克共和国,CZ\n", encoding="utf-8")
    return aliases_path


def _write_web_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "输出表"
    sheet.append(["积分", "乌克兰(USD)", None, "捷克共和国(CZK)", None])
    sheet.append(["积分", "价格", "收入", "价格", "收入"])
    sheet.append([0.99, 0.99, 0.7, 9, 6])
    sheet.append([1.99, 1.99, 1.4, 19, 13])
    sheet["B3"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    workbook.save(path)
    return path
