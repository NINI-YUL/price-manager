"""Read-only Google Play wide-table adapter for P1-004."""

from __future__ import annotations

import hashlib
import re
from collections.abc import Mapping
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from itertools import zip_longest
from pathlib import Path
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from src.models import (
    Channel,
    ImportIssue,
    ImportPreview,
    ImportStatistics,
    ImportTaskStatus,
    IssueSeverity,
    StandardPrice,
)

PRODUCT_TIER_PATTERN = re.compile(r"_iap_(\d+(?:\.\d{1,2})?)(?:&[^\s]*)?$")
COUNTRY_CODE_PATTERN = re.compile(r"^[A-Z]{2}$")
CURRENCY_PATTERN = re.compile(r"^[A-Z]{3}$")
COUNTRY_CODE_HEADERS = frozenset({"国家代码", "countrycode"})
CURRENCY_HEADERS = frozenset({"货币类型", "币种", "currency", "currencycode"})
IGNORED_REGION_HEADERS = frozenset({"国家/地区", "国家地区", "country/region", "region"})
FATAL_READ_ERRORS = (
    OSError,
    InvalidFileException,
    BadZipFile,
    ParseError,
    EOFError,
    ValueError,
)


@dataclass(frozen=True, slots=True)
class GoogleAdapterConfig:
    supported_country_codes: frozenset[str]
    supported_currency_codes: frozenset[str]
    configured_tiers: frozenset[Decimal]
    product_tiers: Mapping[str, Decimal] = field(default_factory=dict)
    country_aliases: Mapping[str, str] = field(default_factory=dict)
    max_file_size_bytes: int = 50 * 1024 * 1024


@dataclass(frozen=True, slots=True)
class _SheetLayout:
    sheet_name: str
    headers: tuple[str, ...]
    country_column: int
    currency_column: int
    ignored_columns: frozenset[int]


@dataclass(frozen=True, slots=True)
class _ProductColumn:
    column_index: int
    column_letter: str
    product_id: str
    usd_tier: Decimal


class GooglePriceAdapter:
    """Parse a Google workbook without mutating it or any database table."""

    def __init__(self, config: GoogleAdapterConfig) -> None:
        self._config = config
        self._supported_country_codes = frozenset(
            value.upper() for value in config.supported_country_codes
        )
        self._supported_currency_codes = frozenset(
            value.upper() for value in config.supported_currency_codes
        )
        self._configured_tiers = frozenset(Decimal(str(value)) for value in config.configured_tiers)
        self._product_tiers = {
            product_id: Decimal(str(tier)) for product_id, tier in config.product_tiers.items()
        }
        self._country_aliases = {
            _normalise_alias(alias): country_code.upper()
            for alias, country_code in config.country_aliases.items()
        }

    def parse(self, file_path: str | Path) -> ImportPreview:
        path = Path(file_path).expanduser().resolve()
        source_path = str(path)
        if path.suffix.lower() != ".xlsx":
            return self._fatal_preview(
                source_path,
                None,
                "G001",
                "Google import supports .xlsx files only",
                source_value=path.suffix or None,
            )
        try:
            file_size = path.stat().st_size
        except OSError as error:
            return self._fatal_preview(
                source_path, None, "G001", f"cannot access source file: {error}"
            )
        if file_size > self._config.max_file_size_bytes:
            return self._fatal_preview(
                source_path,
                None,
                "G001",
                f"source file exceeds {self._config.max_file_size_bytes} bytes",
                source_value=str(file_size),
            )
        try:
            source_sha256 = _sha256(path)
        except OSError as error:
            return self._fatal_preview(
                source_path, None, "G001", f"cannot read source file: {error}"
            )

        value_workbook = None
        formula_workbook = None
        try:
            value_workbook = load_workbook(path, read_only=True, data_only=True)
            formula_workbook = load_workbook(path, read_only=True, data_only=False)
            return self._parse_open_workbooks(
                source_path, source_sha256, value_workbook, formula_workbook
            )
        except FATAL_READ_ERRORS as error:
            return self._fatal_preview(
                source_path,
                source_sha256,
                "G001",
                f"cannot read Google workbook: {error}",
            )
        finally:
            if value_workbook is not None:
                value_workbook.close()
            if formula_workbook is not None:
                formula_workbook.close()

    def _parse_open_workbooks(
        self, source_path: str, source_sha256: str, value_workbook, formula_workbook
    ) -> ImportPreview:
        _repair_read_only_dimensions(value_workbook)
        _repair_read_only_dimensions(formula_workbook)
        layouts = [
            layout
            for sheet in formula_workbook.worksheets
            if (layout := self._detect_layout(sheet)) is not None
        ]
        if len(layouts) != 1:
            message = (
                "no worksheet contains the required country and currency headers"
                if not layouts
                else "multiple worksheets contain the required Google headers"
            )
            return self._fatal_preview(
                source_path,
                source_sha256,
                "G002",
                message,
                source_value=", ".join(layout.sheet_name for layout in layouts) or None,
            )

        layout = layouts[0]
        header_issues = self._validate_headers(layout)
        if header_issues:
            return self._fatal_preview_from_issues(
                source_path, source_sha256, layout.sheet_name, header_issues
            )

        product_columns, issues = self._resolve_product_columns(layout)
        formula_sheet = formula_workbook[layout.sheet_name]
        value_sheet = value_workbook[layout.sheet_name]
        records: list[StandardPrice] = []
        records_by_key: dict[tuple[Channel, str, Decimal, str], StandardPrice] = {}
        currencies_by_country_tier: dict[tuple[str, Decimal], set[str]] = {}
        source_row_count = 0
        price_cell_count = 0
        duplicate_count = 0

        formula_rows = formula_sheet.iter_rows(min_row=2, max_col=len(layout.headers))
        value_rows = value_sheet.iter_rows(min_row=2, max_col=len(layout.headers))
        for source_row, (formula_row, value_row) in enumerate(
            zip_longest(formula_rows, value_rows, fillvalue=()), start=2
        ):
            if _row_is_empty(formula_row, value_row):
                continue
            source_row_count += 1
            price_cell_count += len(product_columns)
            raw_country = _cell_value(value_row, layout.country_column)
            raw_currency = _cell_value(value_row, layout.currency_column)
            country_code, country_issue = self._resolve_country(
                raw_country,
                layout.sheet_name,
                source_row,
                get_column_letter(layout.country_column + 1),
            )
            if country_issue is not None:
                issues.append(country_issue)
            currency, currency_issue = self._resolve_currency(
                raw_currency,
                layout.sheet_name,
                source_row,
                get_column_letter(layout.currency_column + 1),
            )
            if currency_issue is not None:
                issues.append(currency_issue)
            if country_code is None or currency is None:
                continue

            for product in product_columns:
                value_cell = _cell(value_row, product.column_index)
                formula_cell = _cell(formula_row, product.column_index)
                raw_value = value_cell.value if value_cell is not None else None
                formula_value = formula_cell.value if formula_cell is not None else None
                if (
                    formula_cell is not None
                    and formula_cell.data_type == "f"
                    and raw_value is None
                ):
                    issues.append(
                        _issue(
                            "G006",
                            IssueSeverity.ERROR,
                            "formula price has no cached value",
                            layout.sheet_name,
                            source_row,
                            product.column_letter,
                            formula_value,
                        )
                    )
                    continue

                local_price, price_error = _parse_positive_decimal(raw_value)
                if price_error is not None:
                    issues.append(
                        _issue(
                            "G006",
                            IssueSeverity.ERROR,
                            price_error,
                            layout.sheet_name,
                            source_row,
                            product.column_letter,
                            raw_value,
                        )
                    )
                    continue

                record = StandardPrice(
                    channel=Channel.GOOGLE,
                    country_code=country_code,
                    usd_tier=product.usd_tier,
                    currency=currency,
                    local_price=local_price,
                    product_id=product.product_id,
                    source_sheet=layout.sheet_name,
                    source_row=source_row,
                    source_column=product.column_letter,
                )
                country_tier_key = (country_code, product.usd_tier)
                known_currencies = currencies_by_country_tier.setdefault(country_tier_key, set())
                if known_currencies and currency not in known_currencies:
                    issues.append(
                        _issue(
                            "G007",
                            IssueSeverity.ERROR,
                            "the same country and tier use multiple currencies",
                            layout.sheet_name,
                            source_row,
                            product.column_letter,
                            currency,
                        )
                    )
                known_currencies.add(currency)

                existing = records_by_key.get(record.natural_key)
                if existing is not None:
                    if existing.local_price == record.local_price:
                        duplicate_count += 1
                        issues.append(
                            _issue(
                                "G102",
                                IssueSeverity.WARNING,
                                "identical standard price was deduplicated",
                                layout.sheet_name,
                                source_row,
                                product.column_letter,
                                raw_value,
                            )
                        )
                    else:
                        issues.append(
                            _issue(
                                "G007",
                                IssueSeverity.ERROR,
                                "the same standard key has conflicting prices",
                                layout.sheet_name,
                                source_row,
                                product.column_letter,
                                raw_value,
                            )
                        )
                    continue
                records_by_key[record.natural_key] = record
                records.append(record)

        return self._preview(
            source_path=source_path,
            source_sha256=source_sha256,
            selected_sheet=layout.sheet_name,
            records=records,
            issues=issues,
            source_row_count=source_row_count,
            product_count=len(product_columns),
            price_cell_count=price_cell_count,
            duplicate_count=duplicate_count,
        )

    def _detect_layout(self, sheet) -> _SheetLayout | None:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1), ())
        headers = tuple(_clean_header(cell.value) for cell in header_row)
        normalised = tuple(_normalise_header(header) for header in headers)
        country_columns = [
            index for index, header in enumerate(normalised) if header in COUNTRY_CODE_HEADERS
        ]
        currency_columns = [
            index for index, header in enumerate(normalised) if header in CURRENCY_HEADERS
        ]
        if not country_columns or not currency_columns:
            return None
        ignored_columns = frozenset(
            index
            for index, header in enumerate(normalised)
            if header in IGNORED_REGION_HEADERS
        )
        return _SheetLayout(
            sheet_name=sheet.title,
            headers=headers,
            country_column=country_columns[0],
            currency_column=currency_columns[0],
            ignored_columns=ignored_columns,
        )

    def _validate_headers(self, layout: _SheetLayout) -> list[ImportIssue]:
        normalised = tuple(_normalise_header(header) for header in layout.headers)
        issues: list[ImportIssue] = []
        for aliases, label in (
            (COUNTRY_CODE_HEADERS, "country code"),
            (CURRENCY_HEADERS, "currency"),
        ):
            positions = [index for index, header in enumerate(normalised) if header in aliases]
            if len(positions) > 1:
                issues.append(
                    _issue(
                        "G002",
                        IssueSeverity.ERROR,
                        f"duplicate {label} headers",
                        layout.sheet_name,
                        1,
                        get_column_letter(positions[1] + 1),
                        layout.headers[positions[1]],
                    )
                )

        excluded = {
            layout.country_column,
            layout.currency_column,
            *layout.ignored_columns,
        }
        seen_product_headers: dict[str, int] = {}
        for index, header in enumerate(layout.headers):
            if index in excluded or not header:
                continue
            normalised_header = _normalise_header(header)
            if normalised_header in seen_product_headers:
                issues.append(
                    _issue(
                        "G002",
                        IssueSeverity.ERROR,
                        "duplicate product header",
                        layout.sheet_name,
                        1,
                        get_column_letter(index + 1),
                        header,
                    )
                )
            else:
                seen_product_headers[normalised_header] = index
        if not seen_product_headers:
            issues.append(
                _issue(
                    "G002",
                    IssueSeverity.ERROR,
                    "no product headers were found",
                    layout.sheet_name,
                    1,
                )
            )
        return issues

    def _resolve_product_columns(
        self, layout: _SheetLayout
    ) -> tuple[list[_ProductColumn], list[ImportIssue]]:
        products: list[_ProductColumn] = []
        issues: list[ImportIssue] = []
        excluded = {
            layout.country_column,
            layout.currency_column,
            *layout.ignored_columns,
        }
        for index, header in enumerate(layout.headers):
            if index in excluded or not header:
                continue
            tier = self._product_tiers.get(header)
            if tier is None:
                match = PRODUCT_TIER_PATTERN.search(header)
                if match is not None:
                    tier = Decimal(match.group(1))
            if tier is None or tier not in self._configured_tiers:
                issues.append(
                    _issue(
                        "G003",
                        IssueSeverity.ERROR,
                        "product ID cannot be mapped to a configured USD tier",
                        layout.sheet_name,
                        1,
                        get_column_letter(index + 1),
                        header,
                    )
                )
                continue
            products.append(
                _ProductColumn(
                    column_index=index,
                    column_letter=get_column_letter(index + 1),
                    product_id=header,
                    usd_tier=tier,
                )
            )

        present_tiers = {product.usd_tier for product in products}
        for missing_tier in sorted(self._configured_tiers - present_tiers):
            issues.append(
                _issue(
                    "G008",
                    IssueSeverity.ERROR,
                    f"configured tier {missing_tier} has no product column",
                    layout.sheet_name,
                    1,
                    source_value=str(missing_tier),
                )
            )
        return products, issues

    def _resolve_country(
        self,
        raw_value,
        sheet_name: str,
        source_row: int,
        source_column: str,
    ) -> tuple[str | None, ImportIssue | None]:
        text = _clean_value(raw_value)
        candidate = text.upper()
        if COUNTRY_CODE_PATTERN.fullmatch(candidate) and candidate in self._supported_country_codes:
            return candidate, None
        alias = self._country_aliases.get(_normalise_alias(text))
        if alias is not None and alias in self._supported_country_codes:
            return alias, _issue(
                "G101",
                IssueSeverity.WARNING,
                f"country alias was uniquely mapped to {alias}",
                sheet_name,
                source_row,
                source_column,
                raw_value,
            )
        return None, _issue(
            "G004",
            IssueSeverity.ERROR,
            "country code is invalid or unsupported",
            sheet_name,
            source_row,
            source_column,
            raw_value,
        )

    def _resolve_currency(
        self,
        raw_value,
        sheet_name: str,
        source_row: int,
        source_column: str,
    ) -> tuple[str | None, ImportIssue | None]:
        currency = _clean_value(raw_value).upper()
        if (
            CURRENCY_PATTERN.fullmatch(currency)
            and currency in self._supported_currency_codes
        ):
            return currency, None
        return None, _issue(
            "G005",
            IssueSeverity.ERROR,
            "currency is invalid or unsupported",
            sheet_name,
            source_row,
            source_column,
            raw_value,
        )

    def _preview(
        self,
        *,
        source_path: str,
        source_sha256: str,
        selected_sheet: str,
        records: list[StandardPrice],
        issues: list[ImportIssue],
        source_row_count: int,
        product_count: int,
        price_cell_count: int,
        duplicate_count: int,
    ) -> ImportPreview:
        error_count = sum(issue.severity is IssueSeverity.ERROR for issue in issues)
        warning_count = sum(issue.severity is IssueSeverity.WARNING for issue in issues)
        return ImportPreview(
            channel=Channel.GOOGLE,
            source_path=source_path,
            source_sha256=source_sha256,
            selected_sheet=selected_sheet,
            status=ImportTaskStatus.CHECKING,
            records=tuple(records),
            issues=tuple(issues),
            statistics=ImportStatistics(
                source_row_count=source_row_count,
                product_count=product_count,
                price_cell_count=price_cell_count,
                accepted_record_count=len(records),
                country_count=len({record.country_code for record in records}),
                currency_count=len({record.currency for record in records}),
                tier_count=len({record.usd_tier for record in records}),
                duplicate_count=duplicate_count,
                error_count=error_count,
                warning_count=warning_count,
            ),
        )

    def _fatal_preview(
        self,
        source_path: str,
        source_sha256: str | None,
        code: str,
        message: str,
        *,
        source_value: str | None = None,
    ) -> ImportPreview:
        return self._fatal_preview_from_issues(
            source_path,
            source_sha256,
            None,
            [_issue(code, IssueSeverity.ERROR, message, source_value=source_value)],
        )

    def _fatal_preview_from_issues(
        self,
        source_path: str,
        source_sha256: str | None,
        selected_sheet: str | None,
        issues: list[ImportIssue],
    ) -> ImportPreview:
        error_count = sum(issue.severity is IssueSeverity.ERROR for issue in issues)
        warning_count = sum(issue.severity is IssueSeverity.WARNING for issue in issues)
        return ImportPreview(
            channel=Channel.GOOGLE,
            source_path=source_path,
            source_sha256=source_sha256,
            selected_sheet=selected_sheet,
            status=ImportTaskStatus.FAILED,
            records=(),
            issues=tuple(issues),
            statistics=ImportStatistics(
                error_count=error_count,
                warning_count=warning_count,
            ),
        )


def _normalise_header(value: str) -> str:
    return re.sub(r"[\s_]+", "", value).casefold()


def _normalise_alias(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip()).casefold()


def _clean_header(value) -> str:
    return _clean_value(value)


def _clean_value(value) -> str:
    return "" if value is None else str(value).strip()


def _cell(row, index: int):
    return row[index] if index < len(row) else None


def _cell_value(row, index: int):
    cell = _cell(row, index)
    return cell.value if cell is not None else None


def _row_is_empty(formula_row, value_row) -> bool:
    return not any(
        cell is not None and cell.value is not None
        for row in (formula_row, value_row)
        for cell in row
    )


def _parse_positive_decimal(raw_value) -> tuple[Decimal | None, str | None]:
    if raw_value is None or (isinstance(raw_value, str) and not raw_value.strip()):
        return None, "price is blank"
    if isinstance(raw_value, bool):
        return None, "price is not numeric"
    try:
        value = Decimal(str(raw_value).strip())
    except (InvalidOperation, ValueError):
        return None, "price is not numeric"
    if not value.is_finite():
        return None, "price is not finite"
    if value <= 0:
        return None, "price must be greater than zero"
    return value, None


def _issue(
    code: str,
    severity: IssueSeverity,
    message: str,
    sheet_name: str | None = None,
    source_row: int | None = None,
    source_column: str | None = None,
    source_value=None,
) -> ImportIssue:
    return ImportIssue(
        code=code,
        severity=severity,
        message=message,
        sheet_name=sheet_name,
        source_row=source_row,
        source_column=source_column,
        source_value=None if source_value is None else str(source_value),
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _repair_read_only_dimensions(workbook) -> None:
    """Recover rows hidden by incorrect worksheet dimension metadata."""

    for sheet in workbook.worksheets:
        try:
            dimension = sheet.calculate_dimension()
        except ValueError:
            sheet.calculate_dimension(force=True)
            continue
        if dimension == "A1:A1":
            sheet.reset_dimensions()
            sheet.calculate_dimension(force=True)
