"""Read-only third-party web price workbook adapter for P1-006."""

from __future__ import annotations

import hashlib
import re
from collections.abc import Mapping
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from itertools import zip_longest
from pathlib import Path
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from src.models import (
    Channel,
    ImportIssue,
    ImportPreview,
    ImportStatistics,
    ImportTaskStatus,
    IssueSeverity,
    StandardPrice,
)

POINT_HEADERS = frozenset({"积分", "档位", "usd档位", "usdtier"})
PRICE_HEADERS = frozenset({"价格", "price"})
INCOME_HEADERS = frozenset({"收入", "income", "revenue", "proceeds"})
COUNTRY_CURRENCY_PATTERN = re.compile(r"^(?P<country>.+?)[(（](?P<currency>[A-Za-z]{3})[)）]$")
CURRENCY_PATTERN = re.compile(r"^[A-Z]{3}$")
FATAL_READ_ERRORS = (
    OSError,
    InvalidFileException,
    BadZipFile,
    ParseError,
    EOFError,
    ValueError,
)


@dataclass(frozen=True, slots=True)
class WebAdapterConfig:
    country_names: Mapping[str, str]
    supported_currency_codes: frozenset[str]
    configured_tiers: frozenset[Decimal]
    max_file_size_bytes: int = 50 * 1024 * 1024


@dataclass(frozen=True, slots=True)
class _SheetLayout:
    sheet_name: str
    max_column: int


@dataclass(frozen=True, slots=True)
class _CountryPriceColumn:
    country_code: str
    currency: str
    price_index: int
    income_index: int
    price_column: str


class WebPriceAdapter:
    """Parse a web price wide table without changing the source workbook."""

    def __init__(self, config: WebAdapterConfig) -> None:
        self._config = config
        self._country_names = {
            _normalise_country_name(name): code.upper()
            for name, code in config.country_names.items()
        }
        self._supported_currency_codes = frozenset(
            code.upper() for code in config.supported_currency_codes
        )
        self._configured_tiers = frozenset(Decimal(str(tier)) for tier in config.configured_tiers)

    def parse(self, file_path: str | Path) -> ImportPreview:
        path = Path(file_path).expanduser().resolve()
        source_path = str(path)
        if path.suffix.casefold() != ".xlsx":
            return self._fatal_preview(
                source_path,
                None,
                "W001",
                "web price import supports .xlsx files only",
                source_value=path.suffix or None,
            )
        try:
            file_size = path.stat().st_size
        except OSError as error:
            return self._fatal_preview(
                source_path, None, "W001", f"cannot access source file: {error}"
            )
        if file_size > self._config.max_file_size_bytes:
            return self._fatal_preview(
                source_path,
                None,
                "W001",
                f"source file exceeds {self._config.max_file_size_bytes} bytes",
                source_value=str(file_size),
            )
        try:
            source_sha256 = _sha256(path)
        except OSError as error:
            return self._fatal_preview(
                source_path, None, "W001", f"cannot read source file: {error}"
            )

        value_workbook = None
        formula_workbook = None
        try:
            value_workbook = load_workbook(path, read_only=True, data_only=True)
            formula_workbook = load_workbook(path, read_only=True, data_only=False)
            return self._parse_open_workbooks(
                source_path, source_sha256, value_workbook, formula_workbook
            )
        except FATAL_READ_ERRORS as error:
            return self._fatal_preview(
                source_path,
                source_sha256,
                "W001",
                f"cannot read web price workbook: {error}",
            )
        finally:
            if value_workbook is not None:
                value_workbook.close()
            if formula_workbook is not None:
                formula_workbook.close()

    def _parse_open_workbooks(
        self, source_path: str, source_sha256: str, value_workbook, formula_workbook
    ) -> ImportPreview:
        _repair_read_only_dimensions(value_workbook)
        _repair_read_only_dimensions(formula_workbook)
        layouts = [
            layout
            for sheet in formula_workbook.worksheets
            if (layout := self._detect_layout(sheet)) is not None
        ]
        if len(layouts) != 1:
            message = (
                "no worksheet contains the required two-row web price headers"
                if not layouts
                else "multiple worksheets contain the required web price headers"
            )
            return self._fatal_preview(
                source_path,
                source_sha256,
                "W002",
                message,
                source_value=", ".join(layout.sheet_name for layout in layouts) or None,
            )

        layout = layouts[0]
        formula_sheet = formula_workbook[layout.sheet_name]
        value_sheet = value_workbook[layout.sheet_name]
        country_columns, issues = self._resolve_country_columns(formula_sheet, layout)
        structural_issues = [issue for issue in issues if issue.code == "W002"]
        if structural_issues:
            return self._fatal_preview_from_issues(
                source_path, source_sha256, layout.sheet_name, issues
            )

        records: list[StandardPrice] = []
        records_by_key: dict[tuple[Channel, str, Decimal, str], StandardPrice] = {}
        present_tiers: set[Decimal] = set()
        source_row_count = 0
        price_cell_count = 0
        duplicate_count = 0

        formula_rows = formula_sheet.iter_rows(min_row=3, max_col=layout.max_column)
        value_rows = value_sheet.iter_rows(min_row=3, max_col=layout.max_column)
        for source_row, (formula_row, value_row) in enumerate(
            zip_longest(formula_rows, value_rows, fillvalue=()), start=3
        ):
            if _row_is_empty(formula_row, value_row):
                continue
            source_row_count += 1
            price_cell_count += len(country_columns)
            tier_value_cell = _cell(value_row, 0)
            tier_formula_cell = _cell(formula_row, 0)
            raw_tier = tier_value_cell.value if tier_value_cell is not None else None
            formula_tier = tier_formula_cell.value if tier_formula_cell is not None else None
            if (
                tier_formula_cell is not None
                and tier_formula_cell.data_type == "f"
                and raw_tier is None
            ):
                issues.append(
                    _issue(
                        "W005",
                        IssueSeverity.ERROR,
                        "formula tier has no cached value",
                        layout.sheet_name,
                        source_row,
                        "A",
                        formula_tier,
                    )
                )
                continue
            usd_tier, tier_error = _parse_tier(raw_tier, self._configured_tiers)
            if tier_error is not None:
                issues.append(
                    _issue(
                        "W005",
                        IssueSeverity.ERROR,
                        tier_error,
                        layout.sheet_name,
                        source_row,
                        "A",
                        raw_tier,
                    )
                )
                continue
            present_tiers.add(usd_tier)

            for column in country_columns:
                value_cell = _cell(value_row, column.price_index)
                formula_cell = _cell(formula_row, column.price_index)
                raw_price = value_cell.value if value_cell is not None else None
                formula_price = formula_cell.value if formula_cell is not None else None
                if formula_cell is not None and formula_cell.data_type == "f" and raw_price is None:
                    issues.append(
                        _issue(
                            "W006",
                            IssueSeverity.ERROR,
                            "formula price has no cached value",
                            layout.sheet_name,
                            source_row,
                            column.price_column,
                            formula_price,
                        )
                    )
                    continue
                local_price, price_error = _parse_positive_decimal(raw_price)
                if price_error is not None:
                    issues.append(
                        _issue(
                            "W006",
                            IssueSeverity.ERROR,
                            price_error,
                            layout.sheet_name,
                            source_row,
                            column.price_column,
                            raw_price,
                        )
                    )
                    continue

                record = StandardPrice(
                    channel=Channel.WEB,
                    country_code=column.country_code,
                    usd_tier=usd_tier,
                    currency=column.currency,
                    local_price=local_price,
                    product_id=None,
                    source_sheet=layout.sheet_name,
                    source_row=source_row,
                    source_column=column.price_column,
                )
                existing = records_by_key.get(record.natural_key)
                if existing is not None:
                    if existing.local_price == record.local_price:
                        duplicate_count += 1
                        issues.append(
                            _issue(
                                "W102",
                                IssueSeverity.WARNING,
                                "identical web price was deduplicated",
                                layout.sheet_name,
                                source_row,
                                column.price_column,
                                raw_price,
                            )
                        )
                    else:
                        issues.append(
                            _issue(
                                "W007",
                                IssueSeverity.ERROR,
                                "the same web standard key has conflicting prices",
                                layout.sheet_name,
                                source_row,
                                column.price_column,
                                raw_price,
                            )
                        )
                    continue
                records_by_key[record.natural_key] = record
                records.append(record)

        for missing_tier in sorted(self._configured_tiers - present_tiers):
            issues.append(
                _issue(
                    "W008",
                    IssueSeverity.ERROR,
                    f"configured tier {missing_tier} is missing",
                    layout.sheet_name,
                    source_value=str(missing_tier),
                )
            )

        return self._preview(
            source_path=source_path,
            source_sha256=source_sha256,
            selected_sheet=layout.sheet_name,
            records=records,
            issues=issues,
            source_row_count=source_row_count,
            price_cell_count=price_cell_count,
            duplicate_count=duplicate_count,
        )

    def _detect_layout(self, sheet) -> _SheetLayout | None:
        rows = tuple(sheet.iter_rows(min_row=1, max_row=2))
        if len(rows) < 2:
            return None
        first_header = _normalise_header(_cell_value(rows[0], 0))
        second_header = _normalise_header(_cell_value(rows[1], 0))
        second_row = {
            _normalise_header(cell.value) for cell in rows[1][1:] if cell.value is not None
        }
        if (
            first_header not in POINT_HEADERS
            or second_header not in POINT_HEADERS
            or not (second_row & PRICE_HEADERS)
            or not (second_row & INCOME_HEADERS)
        ):
            return None
        return _SheetLayout(sheet_name=sheet.title, max_column=sheet.max_column)

    def _resolve_country_columns(
        self, sheet, layout: _SheetLayout
    ) -> tuple[list[_CountryPriceColumn], list[ImportIssue]]:
        rows = tuple(sheet.iter_rows(min_row=1, max_row=2, max_col=layout.max_column))
        row_one = rows[0] if rows else ()
        row_two = rows[1] if len(rows) > 1 else ()
        issues: list[ImportIssue] = []
        columns: list[_CountryPriceColumn] = []
        if (layout.max_column - 1) % 2:
            issues.append(
                _issue(
                    "W002",
                    IssueSeverity.ERROR,
                    "country price and income columns must appear in pairs",
                    layout.sheet_name,
                    2,
                    get_column_letter(layout.max_column),
                )
            )

        seen_countries: set[str] = set()
        for price_index in range(1, layout.max_column, 2):
            income_index = price_index + 1
            price_column = get_column_letter(price_index + 1)
            raw_header = _clean_value(_cell_value(row_one, price_index))
            adjacent_header = _clean_value(_cell_value(row_one, income_index))
            price_header = _normalise_header(_cell_value(row_two, price_index))
            income_header = _normalise_header(_cell_value(row_two, income_index))
            if (
                not raw_header
                or adjacent_header
                or price_header not in PRICE_HEADERS
                or income_header not in INCOME_HEADERS
            ):
                issues.append(
                    _issue(
                        "W002",
                        IssueSeverity.ERROR,
                        "web country columns require 国家(币种) and 价格/收入 headers",
                        layout.sheet_name,
                        1,
                        price_column,
                        raw_header or adjacent_header or None,
                    )
                )
                continue

            match = COUNTRY_CURRENCY_PATTERN.fullmatch(raw_header)
            if match is None:
                issues.append(
                    _issue(
                        "W002",
                        IssueSeverity.ERROR,
                        "country header must use 国家(币种) format",
                        layout.sheet_name,
                        1,
                        price_column,
                        raw_header,
                    )
                )
                continue

            country_name = match.group("country").strip()
            country_code = self._country_names.get(_normalise_country_name(country_name))
            if country_code is None:
                issues.append(
                    _issue(
                        "W003",
                        IssueSeverity.ERROR,
                        "web country or region name is not configured",
                        layout.sheet_name,
                        1,
                        price_column,
                        country_name,
                    )
                )
                continue

            currency = match.group("currency").upper()
            if (
                not CURRENCY_PATTERN.fullmatch(currency)
                or currency not in self._supported_currency_codes
            ):
                issues.append(
                    _issue(
                        "W004",
                        IssueSeverity.ERROR,
                        "web currency code is invalid or unknown",
                        layout.sheet_name,
                        1,
                        price_column,
                        currency,
                    )
                )
                continue

            if country_code in seen_countries:
                issues.append(
                    _issue(
                        "W007",
                        IssueSeverity.ERROR,
                        "the same web country appears in multiple column pairs",
                        layout.sheet_name,
                        1,
                        price_column,
                        country_name,
                    )
                )
                continue
            seen_countries.add(country_code)
            columns.append(
                _CountryPriceColumn(
                    country_code=country_code,
                    currency=currency,
                    price_index=price_index,
                    income_index=income_index,
                    price_column=price_column,
                )
            )

        if not columns and not issues:
            issues.append(
                _issue(
                    "W002",
                    IssueSeverity.ERROR,
                    "no web country price columns were found",
                    layout.sheet_name,
                    1,
                )
            )
        return columns, issues

    def _preview(
        self,
        *,
        source_path: str,
        source_sha256: str,
        selected_sheet: str,
        records: list[StandardPrice],
        issues: list[ImportIssue],
        source_row_count: int,
        price_cell_count: int,
        duplicate_count: int,
    ) -> ImportPreview:
        error_count = sum(issue.severity is IssueSeverity.ERROR for issue in issues)
        warning_count = sum(issue.severity is IssueSeverity.WARNING for issue in issues)
        return ImportPreview(
            channel=Channel.WEB,
            source_path=source_path,
            source_sha256=source_sha256,
            selected_sheet=selected_sheet,
            status=ImportTaskStatus.CHECKING,
            records=tuple(records),
            issues=tuple(issues),
            statistics=ImportStatistics(
                source_row_count=source_row_count,
                product_count=0,
                price_cell_count=price_cell_count,
                accepted_record_count=len(records),
                country_count=len({record.country_code for record in records}),
                currency_count=len({record.currency for record in records}),
                tier_count=len({record.usd_tier for record in records}),
                duplicate_count=duplicate_count,
                error_count=error_count,
                warning_count=warning_count,
            ),
        )

    def _fatal_preview(
        self,
        source_path: str,
        source_sha256: str | None,
        code: str,
        message: str,
        *,
        source_value: str | None = None,
    ) -> ImportPreview:
        return self._fatal_preview_from_issues(
            source_path,
            source_sha256,
            None,
            [_issue(code, IssueSeverity.ERROR, message, source_value=source_value)],
        )

    def _fatal_preview_from_issues(
        self,
        source_path: str,
        source_sha256: str | None,
        selected_sheet: str | None,
        issues: list[ImportIssue],
    ) -> ImportPreview:
        return ImportPreview(
            channel=Channel.WEB,
            source_path=source_path,
            source_sha256=source_sha256,
            selected_sheet=selected_sheet,
            status=ImportTaskStatus.FAILED,
            records=(),
            issues=tuple(issues),
            statistics=ImportStatistics(
                error_count=sum(issue.severity is IssueSeverity.ERROR for issue in issues),
                warning_count=sum(issue.severity is IssueSeverity.WARNING for issue in issues),
            ),
        )


def _parse_tier(
    raw_value, configured_tiers: frozenset[Decimal]
) -> tuple[Decimal | None, str | None]:
    if raw_value is None or (isinstance(raw_value, str) and not raw_value.strip()):
        return None, "USD tier is blank"
    if isinstance(raw_value, bool):
        return None, "USD tier is not numeric"
    try:
        value = Decimal(str(raw_value).strip())
    except (InvalidOperation, ValueError):
        return None, "USD tier is not numeric"
    if not value.is_finite() or value <= 0:
        return None, "USD tier must be a finite positive number"
    if value not in configured_tiers:
        return None, "USD tier is not configured"
    return value, None


def _parse_positive_decimal(raw_value) -> tuple[Decimal | None, str | None]:
    if raw_value is None or (isinstance(raw_value, str) and not raw_value.strip()):
        return None, "price is blank"
    if isinstance(raw_value, bool):
        return None, "price is not numeric"
    try:
        value = Decimal(str(raw_value).strip())
    except (InvalidOperation, ValueError):
        return None, "price is not numeric"
    if not value.is_finite():
        return None, "price is not finite"
    if value <= 0:
        return None, "price must be greater than zero"
    return value, None


def _normalise_header(value) -> str:
    return re.sub(r"[\s/_-]+", "", _clean_value(value)).casefold()


def _normalise_country_name(value: str) -> str:
    return re.sub(r"\s+", "", value.strip()).casefold()


def _clean_value(value) -> str:
    return "" if value is None else str(value).strip()


def _cell(row, index: int):
    return row[index] if index < len(row) else None


def _cell_value(row, index: int):
    cell = _cell(row, index)
    return cell.value if cell is not None else None


def _row_is_empty(formula_row, value_row) -> bool:
    return not any(
        cell is not None and cell.value is not None
        for row in (formula_row, value_row)
        for cell in row
    )


def _issue(
    code: str,
    severity: IssueSeverity,
    message: str,
    sheet_name: str | None = None,
    source_row: int | None = None,
    source_column: str | None = None,
    source_value=None,
) -> ImportIssue:
    return ImportIssue(
        code=code,
        severity=severity,
        message=message,
        sheet_name=sheet_name,
        source_row=source_row,
        source_column=source_column,
        source_value=None if source_value is None else str(source_value),
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _repair_read_only_dimensions(workbook) -> None:
    """Recover rows hidden by incorrect worksheet dimension metadata."""

    for sheet in workbook.worksheets:
        try:
            dimension = sheet.calculate_dimension()
        except ValueError:
            sheet.calculate_dimension(force=True)
            continue
        if dimension == "A1:A1":
            sheet.reset_dimensions()
            sheet.calculate_dimension(force=True)
